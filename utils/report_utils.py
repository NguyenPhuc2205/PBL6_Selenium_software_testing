"""
Utility for generating Excel test reports
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from typing import List, Dict, Any


class ReportUtils:
    """Test report utility functions"""
    
    @staticmethod
    def create_test_report(file_path: str, test_results: List[Dict[str, Any]]):
        """
        Create Excel test report
        
        Args:
            file_path: Path to save the Excel file
            test_results: List of test result dictionaries with keys:
                - stt: Số thứ tự
                - test_case_id: Mã test case
                - description: Mô tả
                - input: Giá trị đầu vào
                - expected: Kết quả mong đợi
                - actual: Kết quả thực tế
                - result: PASS/FAIL
                - screenshot: Path to screenshot
                - timestamp: Thời gian thực hiện
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "STT",
            "Mã Test Case",
            "Mô tả",
            "Input (Đầu vào)",
            "Expected (Kết quả mong đợi)",
            "Actual (Kết quả thực tế)",
            "Result",
            "Screenshot",
            "Timestamp"
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        column_widths = {
            'A': 8,   # STT
            'B': 18,  # Mã Test Case
            'C': 40,  # Mô tả
            'D': 35,  # Input
            'E': 35,  # Expected
            'F': 35,  # Actual
            'G': 10,  # Result
            'H': 25,  # Screenshot
            'I': 20   # Timestamp
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Write data
        for row_num, result in enumerate(test_results, 2):
            # STT
            cell = ws.cell(row=row_num, column=1, value=result.get('stt', ''))
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = border
            
            # Test Case ID
            cell = ws.cell(row=row_num, column=2, value=result.get('test_case_id', ''))
            cell.alignment = Alignment(horizontal="left", vertical="top")
            cell.border = border
            
            # Description
            cell = ws.cell(row=row_num, column=3, value=result.get('description', ''))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            
            # Input
            cell = ws.cell(row=row_num, column=4, value=result.get('input', ''))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            
            # Expected
            cell = ws.cell(row=row_num, column=5, value=result.get('expected', ''))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            
            # Actual
            cell = ws.cell(row=row_num, column=6, value=result.get('actual', ''))
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            
            # Result (PASS/FAIL)
            test_result = result.get('result', 'FAIL')
            cell = ws.cell(row=row_num, column=7, value=test_result)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.border = border
            
            # Apply color based on result
            if test_result == "PASS":
                cell.fill = pass_fill
                cell.font = Font(bold=True, color="006100")
            else:
                cell.fill = fail_fill
                cell.font = Font(bold=True, color="9C0006")
            
            # Screenshot path
            screenshot_path = result.get('screenshot', '')
            cell = ws.cell(row=row_num, column=8, value=screenshot_path)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border
            
            # Add image if screenshot exists
            if screenshot_path and os.path.exists(screenshot_path):
                try:
                    # Set row height for image
                    ws.row_dimensions[row_num].height = 100
                    
                    # Insert image
                    img = XLImage(screenshot_path)
                    img.width = 150
                    img.height = 100
                    ws.add_image(img, f'H{row_num}')
                except Exception as e:
                    print(f"Could not add image: {e}")
            
            # Timestamp
            cell = ws.cell(row=row_num, column=9, value=result.get('timestamp', ''))
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.border = border
        
        # Add summary
        summary_row = len(test_results) + 3
        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
        
        total = len(test_results)
        passed = sum(1 for r in test_results if r.get('result') == 'PASS')
        failed = total - passed
        pass_rate = (passed / total * 100) if total > 0 else 0
        
        ws.cell(row=summary_row + 1, column=1, value=f"Total: {total}")
        ws.cell(row=summary_row + 2, column=1, value=f"Passed: {passed}").font = Font(color="006100")
        ws.cell(row=summary_row + 3, column=1, value=f"Failed: {failed}").font = Font(color="9C0006")
        ws.cell(row=summary_row + 4, column=1, value=f"Pass Rate: {pass_rate:.2f}%").font = Font(bold=True)
        
        # Save workbook
        wb.save(file_path)
        print(f"Report saved to: {file_path}")
    
    @staticmethod
    def format_input_data(data: dict) -> str:
        """Format input data for display in report"""
        lines = []
        for key, value in data.items():
            if key in ['email', 'password']:
                continue  # Skip credentials
            if isinstance(value, list):
                lines.append(f"{key}: {len(value)} items")
                for i, item in enumerate(value):
                    if isinstance(item, dict):
                        lines.append(f"  [{i}] {item.get('content', '')} - correct: {item.get('is_correct', False)}")
            elif isinstance(value, dict):
                lines.append(f"{key}: {value}")
            else:
                lines.append(f"{key}: {value}")
        return "\n".join(lines)
    
    @staticmethod
    def format_actual_result(success: bool, message: str = "", errors: list = None) -> str:
        """Format actual result for display"""
        if success:
            return f"✓ Success\n{message}"
        else:
            result = f"✗ Failed\n{message}"
            if errors:
                result += f"\nErrors: {', '.join(errors)}"
            return result
