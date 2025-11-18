"""
Utility functions for Excel operations
"""
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any


class ExcelUtils:
    """Excel utility functions"""
    
    @staticmethod
    def read_excel_pandas(file_path: str, sheet_name: str = 0) -> pd.DataFrame:
        """Read Excel file using pandas"""
        return pd.read_excel(file_path, sheet_name=sheet_name)
    
    @staticmethod
    def read_excel_to_dict(file_path: str, sheet_name: str = 0) -> List[Dict[str, Any]]:
        """Read Excel file and convert to list of dictionaries"""
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df.to_dict('records')
    
    @staticmethod
    def write_excel_pandas(file_path: str, data: pd.DataFrame, sheet_name: str = 'Sheet1') -> None:
        """Write DataFrame to Excel file"""
        data.to_excel(file_path, sheet_name=sheet_name, index=False)
    
    @staticmethod
    def write_excel_from_dict(file_path: str, data: List[Dict[str, Any]], sheet_name: str = 'Sheet1') -> None:
        """Write list of dictionaries to Excel file"""
        df = pd.DataFrame(data)
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
    
    @staticmethod
    def read_excel_openpyxl(file_path: str, sheet_name: str = None) -> List[List[Any]]:
        """Read Excel file using openpyxl"""
        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        
        return data
    
    @staticmethod
    def write_excel_styled(file_path: str, data: List[List[Any]], headers: List[str] = None) -> None:
        """Write Excel file with styling using openpyxl"""
        wb = Workbook()
        ws = wb.active
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers if provided
        if headers:
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            start_row = 2
        else:
            start_row = 1
        
        # Write data
        for row_num, row_data in enumerate(data, start_row):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    @staticmethod
    def append_to_excel(file_path: str, data: List[List[Any]], sheet_name: str = None) -> None:
        """Append data to existing Excel file"""
        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        for row_data in data:
            ws.append(row_data)
        
        wb.save(file_path)
